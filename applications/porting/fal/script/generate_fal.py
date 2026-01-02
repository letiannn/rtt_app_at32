import os
import shutil
import argparse
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

# python.exe .\generate_fal.py .\fal.xlsx

# 定义输出目录
output_dir = 'out'
inc_dir = '../'

# 自定义过滤器：rjust
def rjust_filter(value, width):
    return str(value).rjust(width)

def read_partition_table(file_path):
    """读取Excel中的分区表并打印出来"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # 存储分区数据
    ram_partitions = []  
    onchip_partitions = []
    nor_partitions = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            size_kb = float(row[4]) if row[4] is not None else 0
            start_addr_decimal = row[2]
            start_addr_hex = row[3]

            partition = {
                'dev': row[0],
                'name': row[1],
                'start_addr': start_addr_decimal,
                'start_hex': start_addr_hex,
                'size_kb': size_kb,
                'size_bytes': int(size_kb*1024)
            }

            if row[0] == "ram_flash": 
                ram_partitions.append(partition)
            elif row[0] == "onchip_flash":
                onchip_partitions.append(partition)
            elif row[0] == "w25q64":
                nor_partitions.append(partition)

        except (ValueError, TypeError) as e:
            print(f"Skipping invalid row: {row}. Error: {e}")

    return ram_partitions, onchip_partitions, nor_partitions

def generate_cfg_file(ram_partitions, onchip_partitions, nor_partitions, template_path):
    """Render the template and generate the fal_cfg.h file"""
    env = Environment(
        loader=FileSystemLoader(template_path),
        trim_blocks=True,
        lstrip_blocks=True
    )
    env.filters['rjust'] = rjust_filter

    template = env.get_template('fal_cfg_template.h.j2')

    # 获取设备名称
    ram_device = ram_partitions[0]['dev'] if ram_partitions else 'ram_flash' 
    onchip_device = onchip_partitions[0]['dev'] if onchip_partitions else 'onchip_flash'
    nor_device = nor_partitions[0]['dev'] if nor_partitions else 'w25q64'

    data = {
        'ram_device_macro': ram_device, 
        'onchip_device_macro': onchip_device,
        'nor_device_macro': nor_device,
        'ram_partitions': ram_partitions,  
        'onchip_partitions': onchip_partitions,
        'nor_partitions': nor_partitions,
    }

    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 生成输出路径
    output_path = os.path.join(output_dir, 'fal_cfg.h')

    with open(output_path, 'w') as f:
        f.write(template.render(data))

    # 复制到 inc 目录
    if not os.path.exists(inc_dir):
        os.makedirs(inc_dir)
    shutil.copy(output_path, os.path.join(inc_dir, 'fal_cfg.h'))

    print(f"Generated configuration file at {output_path}")

# 主程序
if __name__ == "__main__":
    template_path = 'templates'  # 模板目录

    parser = argparse.ArgumentParser(description="生成 FlashDB 和 RamDB 代码")
    parser.add_argument('excel_file', help="Excel 文件路径，例如 fdb.xlsx")
    args = parser.parse_args()
    excel_file_path = args.excel_file

    if not os.path.exists(excel_file_path):
        print(f"错误: 文件 '{excel_file_path}' 不存在！")
        exit(1)

    # 读取数据并生成文件
    ram_partitions, onchip_partitions, nor_partitions = read_partition_table(excel_file_path)
    generate_cfg_file(ram_partitions, onchip_partitions, nor_partitions, template_path)